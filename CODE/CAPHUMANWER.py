import openpyxl

# Function to calculate Word Error Rate (WER)
def wer(ref, hyp):
    # Create a matrix to store distances
    d = [[0] * (len(hyp) + 1) for _ in range(len(ref) + 1)]
    # Initialize first row and column of the matrix
    for i in range(len(ref) + 1):
        d[i][0] = i
    for j in range(len(hyp) + 1):
        d[0][j] = j
    # Populate the matrix
    for i in range(1, len(ref) + 1):
        for j in range(1, len(hyp) + 1):
            cost = 0 if ref[i - 1] == hyp[j - 1] else 1
            d[i][j] = min(d[i - 1][j] + 1, d[i][j - 1] + 1, d[i - 1][j - 1] + cost)
    # Calculate WER
    wer_value = float(d[len(ref)][len(hyp)]) / len(ref)
    return wer_value * 100  # Convert WER to percentage

# Load the Excel workbook
workbook = openpyxl.load_workbook("C:/AudioCAPTCHASamples/VSCODEWER.xlsx")

# Select the active sheet
sheet = workbook.active

# Get the total number of rows in the sheet
total_rows = sheet.max_row

# Create a new workbook
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active

# Set the header for the new sheet
new_sheet['A1'] = "Human Transcription"
new_sheet['B1'] = "Ground Truth"
new_sheet['C1'] = "WER (%)"

# Iterate over rows to calculate WER and create new sheet
for row in range(2, total_rows + 1):  # Start from row 2 to skip header
    ground_truth = str(sheet.cell(row=row, column=1).value)
    human_transcribed = str(sheet.cell(row=row, column=2).value)
    
    # Convert human transcriptions to lowercase
    human_transcribed_lower = human_transcribed.lower()
    
    # Calculate WER
    wer_percent = wer(ground_truth, human_transcribed_lower)
    
    # Write data to the new sheet
    new_sheet.cell(row=row, column=1).value = human_transcribed_lower
    new_sheet.cell(row=row, column=2).value = ground_truth
    new_sheet.cell(row=row, column=3).value = wer_percent

# Save the new workbook for CAP
new_workbook.save("C:/AudioCAPTCHASamples/CAPHWER.xlsx")
